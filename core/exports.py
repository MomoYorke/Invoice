# -*- coding: utf-8 -*-
"""
Pacchetto per la commercialista:
- Excel con registro fatture, riepilogo mensile/trimestrale e per cliente
- PDF di riepilogo
- copia dei PDF delle fatture dell'anno
- tutto in una cartella Esporti/Commercialista_YYYY + zip
"""
import os
import re
import shutil
import zipfile
import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer)
from reportlab.lib.styles import getSampleStyleSheet

from . import stats
from . import lingua as L
from .money import fmt_chf

APP_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXPORT_DIR = os.path.join(APP_DIR, 'Esporti')

HDR_FILL = PatternFill('solid', fgColor='1F4E5F')
HDR_FONT = Font(bold=True, color='FFFFFF')
BOLD = Font(bold=True)
THIN = Border(*[Side(style='thin', color='CCCCCC')] * 4)
NUMFMT = "#,##0.00"


def _sheet_header(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(i)].width = w


def build_excel(con, year, out_path, settings, lingua=None):
    def t(frase, **valori):
        """La frase nella lingua dell'app, coi buchi gia' riempiti."""
        testo = L.t(frase, lingua)
        return testo.format(**valori) if valori else testo

    mesi = L.mesi_elenco(lingua)
    oggi = datetime.date.today().strftime('%d.%m.%Y')
    wb = openpyxl.Workbook()

    # ---- Registro fatture ----
    ws = wb.active
    ws.title = t('Registro fatture')
    ws['A1'] = f"{settings['business_name']} — " + t('Registro fatture {anno}', anno=year)
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = t('Generato il {data} — importi in CHF', data=oggi)
    ws['A2'].font = Font(italic=True, size=9)
    _sheet_header(ws, 4, [t('Nr.'), t('Data'), t('Cliente'), t('Descrizione'),
                          t('Importo CHF'), t('Stato')],
                  [8, 12, 26, 52, 14, 12])
    invs = con.execute(
        'SELECT * FROM invoices WHERE year=? AND deleted_at IS NULL ORDER BY COALESCE(number, 0), date', (year,)).fetchall()
    r = 5
    first_data = r
    for inv in invs:
        descs = [i['description'] for i in con.execute(
            'SELECT description FROM items WHERE invoice_id=? ORDER BY pos', (inv['id'],)) if i['description']]
        date_h = ''
        if inv['date']:
            d = datetime.date.fromisoformat(inv['date'])
            date_h = d.strftime('%d.%m.%Y')
        ws.cell(row=r, column=1, value=f"#{inv['number']}" if inv['number'] else '—')
        ws.cell(row=r, column=2, value=date_h)
        ws.cell(row=r, column=3, value=inv['client_name'])
        ws.cell(row=r, column=4, value='; '.join(descs))
        c = ws.cell(row=r, column=5,
                    value=(inv['total_cents'] / 100) if inv['total_cents'] is not None else None)
        c.number_format = NUMFMT
        if inv['total_cents'] is None:
            ws.cell(row=r, column=4).value = ((ws.cell(row=r, column=4).value or '')
                                              + ' [' + t('importo da verificare') + ']').strip()
        ws.cell(row=r, column=6, value=t(inv['status']))
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = THIN
        r += 1
    ws.cell(row=r, column=4, value=t('TOTALE')).font = BOLD
    tc = ws.cell(row=r, column=5, value=f'=SUM(E{first_data}:E{r - 1})')
    tc.number_format = NUMFMT
    tc.font = BOLD
    total_exact = sum(i['total_cents'] or 0 for i in invs)
    ws.cell(row=r + 1, column=4,
            value=t('Controllo (calcolo interno esatto):')).font = Font(italic=True, size=9)
    ws.cell(row=r + 1, column=5, value=fmt_chf(total_exact)).font = Font(italic=True, size=9)

    # ---- Riepilogo mensile / trimestrale ----
    ws2 = wb.create_sheet(t('Riepilogo mensile'))
    ws2['A1'] = t('Riepilogo mensile {anno}', anno=year)
    ws2['A1'].font = Font(bold=True, size=14)
    _sheet_header(ws2, 3, [t('Mese'), t('Fatturato CHF'), t('N. fatture')], [16, 16, 12])
    months = stats.monthly(con, year)
    counts = [0] * 12
    for inv in invs:
        if inv['date']:
            counts[int(inv['date'][5:7]) - 1] += 1
    rr = 4
    for m in range(12):
        ws2.cell(row=rr, column=1, value=mesi[m])
        c = ws2.cell(row=rr, column=2, value=months[m] / 100)
        c.number_format = NUMFMT
        ws2.cell(row=rr, column=3, value=counts[m])
        rr += 1
    ws2.cell(row=rr, column=1, value=t('TOTALE')).font = BOLD
    c = ws2.cell(row=rr, column=2, value=f'=SUM(B4:B{rr - 1})')
    c.number_format = NUMFMT
    c.font = BOLD
    ws2.cell(row=rr, column=3, value=f'=SUM(C4:C{rr - 1})').font = BOLD
    rr += 2
    _sheet_header(ws2, rr, [t('Trimestre'), t('Fatturato CHF'), ''], [16, 16, 12])
    rr += 1
    for q in range(4):
        tot_q = sum(months[q * 3:q * 3 + 3])
        ws2.cell(row=rr, column=1,
                 value=f'Q{q + 1} ({mesi[q * 3][:3]}–{mesi[q * 3 + 2][:3]})')
        c = ws2.cell(row=rr, column=2, value=tot_q / 100)
        c.number_format = NUMFMT
        rr += 1

    # ---- Per cliente ----
    ws3 = wb.create_sheet(t('Per cliente'))
    ws3['A1'] = t('Fatturato per cliente {anno}', anno=year)
    ws3['A1'].font = Font(bold=True, size=14)
    _sheet_header(ws3, 3, [t('Cliente'), t('Fatturato CHF'), t('N. fatture')], [30, 16, 12])
    rr = 4
    for name, tot, n in stats.by_client(con, year):
        ws3.cell(row=rr, column=1, value=name)
        c = ws3.cell(row=rr, column=2, value=tot / 100)
        c.number_format = NUMFMT
        ws3.cell(row=rr, column=3, value=n)
        rr += 1
    ws3.cell(row=rr, column=1, value=t('TOTALE')).font = BOLD
    c = ws3.cell(row=rr, column=2, value=f'=SUM(B4:B{rr - 1})')
    c.number_format = NUMFMT
    c.font = BOLD

    wb.save(out_path)
    return out_path


def build_summary_pdf(con, year, out_path, settings, lingua=None):
    def t(frase, **valori):
        """La frase nella lingua dell'app, coi buchi gia' riempiti."""
        testo = L.t(frase, lingua)
        return testo.format(**valori) if valori else testo

    mesi = L.mesi_elenco(lingua)
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(out_path, pagesize=A4,
                            leftMargin=2 * cm, rightMargin=2 * cm,
                            topMargin=2 * cm, bottomMargin=2 * cm)
    story = []
    story.append(Paragraph(f"{settings['business_name']}", styles['Title']))
    story.append(Paragraph(
        t('Riepilogo fatture {anno} — per {nome} ({citta})', anno=year,
          nome=settings['accountant_name'], citta=settings['accountant_city']),
        styles['Heading2']))
    story.append(Paragraph(
        t('Generato il {data} — UID {uid} — IBAN {iban}',
          data=datetime.date.today().strftime('%d.%m.%Y'),
          uid=settings['business_uid'], iban=settings['business_iban']),
        styles['Normal']))
    story.append(Spacer(1, 14))

    invs = con.execute('SELECT * FROM invoices WHERE year=? AND deleted_at IS NULL ORDER BY COALESCE(number,0), date',
                       (year,)).fetchall()
    data = [[t('Nr.'), t('Data'), t('Cliente'), t('Importo'), t('Stato')]]
    total = 0
    for inv in invs:
        d = ''
        if inv['date']:
            d = datetime.date.fromisoformat(inv['date']).strftime('%d.%m.%Y')
        data.append([f"#{inv['number']}" if inv['number'] else '—', d,
                     inv['client_name'], fmt_chf(inv['total_cents']), t(inv['status'])])
        total += inv['total_cents'] or 0
    data.append(['', '', t('TOTALE'), fmt_chf(total), ''])
    tabella = Table(data, colWidths=[1.6 * cm, 2.6 * cm, 6.4 * cm, 3.6 * cm, 2.4 * cm], repeatRows=1)
    tabella.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 9),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E5F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONT', (0, 1), (-1, -1), 'Helvetica', 9),
        ('FONT', (0, -1), (-1, -1), 'Helvetica-Bold', 9),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#BBBBBB')),
        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F3F6F8')]),
    ]))
    story.append(tabella)
    story.append(Spacer(1, 16))

    months = stats.monthly(con, year)
    md = [[t('Mese'), t('Fatturato')]] + [[mesi[m], fmt_chf(months[m])] for m in range(12)]
    md.append([t('TOTALE'), fmt_chf(sum(months))])
    tabella_mesi = Table(md, colWidths=[4 * cm, 4 * cm])
    tabella_mesi.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 9),
        ('FONT', (0, 1), (-1, -1), 'Helvetica', 9),
        ('FONT', (0, -1), (-1, -1), 'Helvetica-Bold', 9),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#BBBBBB')),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
    ]))
    story.append(Paragraph(t('Riepilogo mensile'), styles['Heading3']))
    story.append(tabella_mesi)
    doc.build(story)
    return out_path


# --- il PDF che sta accanto al Word ------------------------------------------
# Nell'archivio storico lo stesso documento porta due nomi diversi: «Sofia
# #58.docx» e «Sofia ^N58.pdf». Cambia solo il segno davanti al numero (#, ^N,
# ^LN, ^, _, uno spazio): il nome davanti e il numero in fondo restano quelli.
_CODA_NUMERO = re.compile(r'[\s_#]*(?:\^[A-Za-z]*)?[\s_#]*(\d+)\s*$')


def _nome_e_numero(percorso):
    """Il nome e il numero scritti nel nome del file, separati."""
    base = os.path.splitext(os.path.basename(percorso))[0]
    m = _CODA_NUMERO.search(base)
    if not m:
        return ' '.join(base.split()).lower(), None
    return ' '.join(base[:m.start()].split()).lower(), int(m.group(1))


def pdf_gemello(percorso):
    """Il PDF dello stesso documento, anche se i due nomi non coincidono.

    Si cerca il numero SCRITTO NEL NOME DEL FILE, non il numero della fattura:
    una fattura rinumerata (la #58 diventata #59) sul disco porta ancora il
    vecchio nome, e cercare «59» non troverebbe niente.

    A parita' di numero vince chi ha lo stesso nome davanti, perche' proprio
    una rinumerazione lascia due «58» di due persone diverse nella stessa
    cartella: mandare alla commercialista la fattura di Tizio col numero di
    Caio sarebbe molto peggio che non mandarne nessuna. Se il nome non
    combacia con nessuno, il PDF va bene solo se e' l'unico con quel numero.
    """
    base = os.path.splitext(percorso)[0]
    if os.path.exists(base + '.pdf'):
        return base + '.pdf'
    cartella = os.path.dirname(percorso)
    if not os.path.isdir(cartella):
        return None
    nome, numero = _nome_e_numero(percorso)
    if numero is None:
        return None
    ripiego = []
    for f in sorted(os.listdir(cartella)):
        if not f.lower().endswith('.pdf'):
            continue
        altro_nome, altro_numero = _nome_e_numero(f)
        if altro_numero != numero:
            continue
        if altro_nome == nome:
            return os.path.join(cartella, f)
        ripiego.append(os.path.join(cartella, f))
    return ripiego[0] if len(ripiego) == 1 else None


def _nome_copia(inv, src):
    """Come si chiama la copia del PDF dentro il pacchetto.

    Sul disco i file storici portano nomi di ogni tipo, e a volte un numero
    che non e' piu' quello della fattura: la #58 rinumerata #59 si chiama
    ancora «Sofia ^N58.pdf». Dentro il pacchetto il nome deve combaciare col
    registro, se no chi lo riceve non sa quale riga sta guardando.
    """
    if not inv['number']:
        return os.path.basename(src)
    nome = re.sub(r'[\\/:]', '-', (inv['client_name'] or '').strip())
    return ('#%s %s' % (inv['number'], nome)).strip() + '.pdf'


def lingua_pacchetto(settings, lingua_app):
    """In che lingua esce il pacchetto.

    Non e' la lingua dell'app: il registro lo legge chi tiene la contabilita'.
    Di solito le due coincidono e non c'e' niente da scegliere; quando non
    coincidono vince quella scritta apposta.
    """
    return settings.get('accountant_lingua') or lingua_app


def build_package(con, year, settings, source_root, lingua=None):
    """Crea la cartella Esporti/Commercialista_YYYY con Excel, PDF riepilogo,
    copie delle fatture PDF e uno zip pronto da mandare."""
    stamp = datetime.date.today().strftime('%Y%m%d')
    folder = os.path.join(EXPORT_DIR, f'Commercialista_{year}')
    if os.path.exists(folder):
        shutil.rmtree(folder)
    inv_dir = os.path.join(folder, 'Fatture PDF')
    os.makedirs(inv_dir, exist_ok=True)

    xlsx = build_excel(con, year, os.path.join(folder, f'Registro_Fatture_{year}.xlsx'),
                       settings, lingua)
    spdf = build_summary_pdf(con, year, os.path.join(folder, f'Riepilogo_{year}.pdf'),
                             settings, lingua)

    # copia PDF fatture: prima quelle dell'app, poi quelle importate dalla cartella storica
    copied, missing = 0, []
    for inv in con.execute('SELECT * FROM invoices WHERE year=? AND deleted_at IS NULL ORDER BY COALESCE(number,0)', (year,)):
        src = None
        if inv['pdf_path'] and os.path.exists(inv['pdf_path']):
            src = inv['pdf_path']
        elif inv['source_file']:
            cand = os.path.join(source_root, inv['source_file'])
            if cand.lower().endswith('.pdf') and os.path.exists(cand):
                src = cand
            else:
                src = pdf_gemello(cand)
        if src:
            shutil.copy2(src, os.path.join(inv_dir, _nome_copia(inv, src)))
            copied += 1
        else:
            missing.append(f"#{inv['number'] or '—'} {inv['client_name']}")

    zpath = os.path.join(EXPORT_DIR, f'Commercialista_{settings["accountant_name"].replace(" ", "_")}_{year}_{stamp}.zip')
    with zipfile.ZipFile(zpath, 'w', zipfile.ZIP_DEFLATED) as z:
        for base, _, files in os.walk(folder):
            for f in files:
                p = os.path.join(base, f)
                z.write(p, os.path.relpath(p, os.path.dirname(folder)))
    return {'folder': folder, 'xlsx': xlsx, 'pdf': spdf, 'zip': zpath,
            'copied': copied, 'missing': missing}
